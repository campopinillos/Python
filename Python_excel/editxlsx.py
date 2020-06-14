"""File to work excel files with pandas"""
import pandas as pd
import numpy as np
import matplotlib.pyplot
import openpyxl
import xlsxwriter


# Generando Cálculos con Pandas
df = pd.read_excel(io='informal-data.xlsx', sheet_name='InfoInicial', usecols='A:H', parse_dates=True)
df.head()
df.columns

# 1. Cual fue el usuario que más dinero generó para la empresa, y cual fue el detalle de sus compras?
# Usuario
user = df.groupby(['Usuario'])['Total'].sum().sort_values(ascending=False).head(1)
text = "df.groupby(['Usuario'])['Total'].sum().sort_values(ascending=False).head(1)"
print(type(user))
user1 = df.groupby(['Usuario']).agg('sum').sort_values('Total', ascending=False).head(1)
text1 = "df.groupby(['Usuario']).agg('sum').sort_values('Total', ascending=False).head(1)"
# Detalle
detalle = df[df.Usuario == user1.index[0]]
text2 = 'df[df.Usuario==user1.index[0]]'

# 2. Cual fue el mes donde más dinero se generó y cual fue el detalle de todos los planes?
print(type(df['Fecha']))
df['Mes'] = pd.to_datetime(df['Fecha']).dt.month
mes = df.groupby(['Mes']).agg('sum').sort_values('Total', ascending=False).head(1)
text3 = "df.groupby(['Mes']).agg('sum').sort_values('Total', ascending=False).head(1)"
# Detalle por Planes
detalle1 = df[df.Mes == mes.index[0]].groupby(['Mes', 'Plan']).agg('sum').sort_values('Total', ascending=False).drop(columns=['Usuario'])

# 3. Cuantas transacciones se realizaron con pagos de funciones adicionales y cuanto fue el valor recaudado por los adicionales.
transaciones = df['Adicionales'][df.Adicionales >= 1].count()
AdicionalesVal = df['Valor Adicionales'].sum()
numpy_data = np.array([[transaciones], [AdicionalesVal]])
result = pd.DataFrame({'Transac Adicionales': numpy_data[0], 'Valor Adicionales': numpy_data[1]})

# 4. Realice una gráfica donde se vea como se comportaron las ventas a través de los diferentes meses del año. Con la información resultante proponga opciones de mejora para los meses de peor desempeño.
ventas_mes = df.groupby(['Mes']).agg('sum').sort_index(axis=0).drop(columns=['Usuario'])



# Guardas en Excel con Pandas y openpyxl https://xlsxwriter.readthedocs.io/working_with_pandas.html
# https://stackoverflow.com/questions/28142420/can-pandas-read-and-modify-a-single-excel-file-worksheet-tab-without-modifying

# xlsxwriter First Method does not work the Append function
writer = pd.ExcelWriter('informal-data.xlsx', engine='xlsxwriter', mode='A')

# Punto 1
user.to_excel(writer, sheet_name='Hoja1')
user1.to_excel(writer, sheet_name='Hoja1', startcol=3)
worksheet = writer.sheets['Hoja1']
worksheet.write(2, 0, text)
detalle.to_excel(writer, sheet_name='Hoja1', startrow=4)


# Punto 2
mes.to_excel(writer, sheet_name='Hoja2')
worksheet = writer.sheets['Hoja2']
detalle1.to_excel(writer, sheet_name='Hoja2', startrow=4)

# Punto 3
result.to_excel(writer, sheet_name='Hoja3')

# Punto 4
ventas_mes.to_excel(writer, sheet_name='Hoja4')
# Get the xlsxwriter workbook and worksheet objects.
workbook = writer.book
worksheet = writer.sheets['Hoja4']
# Create a chart object.
chart = workbook.add_chart({'type': 'column'})
# Configure the series of the chart from the dataframe data.
chart.add_series({'values': '=Hoja4!$E$2:$E$13'})
# Insert the chart into the worksheet.
worksheet.insert_chart('G2', chart)

writer.save()


# openpyxl Second Method does work the Append

writer = pd.ExcelWriter('informal-data.xlsx', engine='openpyxl', mode='a')
user.to_excel(writer, sheet_name='Hoja1')
user1.to_excel(writer, sheet_name='Hoja1', startcol=3)
worksheet = writer.sheets['Hoja1']
worksheet.cell(row=3, column=1).value = text
worksheet.cell(row=3, column=4).value = text1

detalle.to_excel(writer, sheet_name='Hoja1', startrow=4)

# Punto 2
mes.to_excel(writer, sheet_name='Hoja2')
worksheet = writer.sheets['Hoja2']
detalle1.to_excel(writer, sheet_name='Hoja2', startrow=4)

# Punto 3
result.to_excel(writer, sheet_name='Hoja3')

# Punto 4 para el Chart https://openpyxl.readthedocs.io/en/stable/charts/bar.ht
ventas_mes.to_excel(writer, sheet_name='Hoja4')
writer.save()


# It does work in safe mode also
with pd.ExcelWriter('informal-data.xlsx', engine='openpyxl', mode='a') as writer:
    user.to_excel(writer, sheet_name='Hoja1')
    user1.to_excel(writer, sheet_name='Hoja1', startcol=3)
    worksheet = writer.sheets['Hoja1']
    worksheet.cell(row=3, column=1).value = text
    worksheet.cell(row=3, column=4).value = text1
